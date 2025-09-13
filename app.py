from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, session, send_from_directory, abort, Response
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os, sqlite3, csv, zipfile
from contextlib import closing
from io import StringIO, BytesIO

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}
MAX_FILES_PER_RECORD = 6

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")
max_len_mb = int(os.environ.get("MAX_CONTENT_LENGTH_MB", "20"))
app.config["MAX_CONTENT_LENGTH"] = max_len_mb * 1024 * 1024

DB_PATH = os.path.join(BASE_DIR, "app.db")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with closing(get_db()) as db:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                is_admin INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                device_name TEXT NOT NULL,
                fusion_count INTEGER NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS photos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_id INTEGER NOT NULL,
                filename TEXT NOT NULL,
                FOREIGN KEY(record_id) REFERENCES records(id)
            );
        """)
        cols = db.execute("PRAGMA table_info(users)").fetchall()
        colnames = {c[1] for c in cols}
        if "is_admin" not in colnames:
            db.execute("ALTER TABLE users ADD COLUMN is_admin INTEGER NOT NULL DEFAULT 0;")
        db.commit()

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

@app.before_request
def ensure_db_initialized():
    if not hasattr(app, "_db_initialized"):
        init_db()
        app._db_initialized = True

def login_required(view):
    from functools import wraps
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped

def admin_required(view):
    from functools import wraps
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if not session.get("is_admin"):
            flash("Acesso restrito ao administrador.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped

@app.route("/register", methods=["GET", "POST"])
def register():
    db = get_db()
    total_users = db.execute("SELECT COUNT(1) FROM users").fetchone()[0]
    if total_users > 0:
        flash("Registro desativado. Apenas o administrador pode criar novos usuários.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if not username or not password:
            flash("Preencha todos os campos.", "error")
            return redirect(url_for("register"))
        is_admin = 1
        pw_hash = generate_password_hash(password)
        try:
            db.execute("INSERT INTO users (username, password_hash, is_admin) VALUES (?, ?, ?)", (username, pw_hash, is_admin))
            db.commit()
            flash("Usuário criado. Faça login.", "success")
            return redirect(url_for("login"))
        except sqlite3.IntegrityError:
            flash("Nome de usuário já existe.", "error")
            return redirect(url_for("register"))
    return render_template("register.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        db = get_db()
        row = db.execute("SELECT id, password_hash, is_admin FROM users WHERE username = ?", (username,)).fetchone()
        if not row or not check_password_hash(row["password_hash"], password):
            flash("Credenciais inválidas.", "error")
            return redirect(url_for("login"))
        session["user_id"] = row["id"]
        session["username"] = username
        session["is_admin"] = bool(row["is_admin"])
        return redirect(url_for("dashboard"))
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Sessão encerrada.", "info")
    return redirect(url_for("login"))

@app.route("/")
@login_required
def dashboard():
    db = get_db()
    recs = db.execute(
        "SELECT r.id, r.device_name, r.fusion_count, r.created_at FROM records r WHERE r.user_id = ? ORDER BY r.created_at DESC",
        (session["user_id"],),
    ).fetchall()
    return render_template("dashboard.html", records=recs)

@app.route("/new", methods=["GET", "POST"])
@login_required
def new_record():
    if request.method == "POST":
        device_name = request.form.get("device_name", "").strip()
        fusion_count = request.form.get("fusion_count", "").strip()
        files = request.files.getlist("photos")
        if not device_name or not fusion_count:
            flash("Informe nome do dispositivo e número de fusões.", "error")
            return redirect(url_for("new_record"))
        try:
            fusion_count = int(fusion_count)
            if fusion_count < 0:
                raise ValueError
        except ValueError:
            flash("Número de fusões inválido.", "error")
            return redirect(url_for("new_record"))
        if len(files) > MAX_FILES_PER_RECORD:
            flash(f"Máximo de {MAX_FILES_PER_RECORD} fotos por registro.", "error")
            return redirect(url_for("new_record"))
        db = get_db()
        cur = db.execute("INSERT INTO records (user_id, device_name, fusion_count) VALUES (?, ?, ?)", (session["user_id"], device_name, fusion_count))
        record_id = cur.lastrowid
        saved_any = False
        for file in files:
            if not file or file.filename == "":
                continue
            if not allowed_file(file.filename):
                flash("Tipo de arquivo não permitido (png/jpg/jpeg/gif/webp).", "error")
                continue
            fname = secure_filename(file.filename)
            base, ext = os.path.splitext(fname)
            final_name = f"{base}{ext}"
            cnt = 1
            while os.path.exists(os.path.join(app.config["UPLOAD_FOLDER"], final_name)):
                final_name = f"{base}_{cnt}{ext}"; cnt += 1
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], final_name))
            db.execute("INSERT INTO photos (record_id, filename) VALUES (?, ?)", (record_id, final_name))
            saved_any = True
        db.commit()
        if not saved_any and len(files) > 0:
            flash("Nenhuma foto foi salva (verifique os tipos permitidos).", "warning")
        else:
            flash("Registro criado com sucesso!", "success")
        return redirect(url_for("dashboard"))
    return render_template("new_record.html")

@app.route("/record/<int:record_id>")
@login_required
def view_record(record_id):
    db = get_db()
    rec = db.execute("SELECT id, device_name, fusion_count, created_at FROM records WHERE id = ? AND user_id = ?", (record_id, session["user_id"])).fetchone()
    if not rec:
        abort(404)
    photos = db.execute("SELECT id, filename FROM photos WHERE record_id = ?", (record_id,)).fetchall()
    return render_template("view_record.html", rec=rec, photos=photos)

@app.route("/uploads/<path:filename>")
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

@app.route("/record/<int:record_id>/delete", methods=["POST"])
@login_required
def delete_record(record_id):
    db = get_db()
    rec = db.execute("SELECT id FROM records WHERE id = ? AND user_id = ?", (record_id, session["user_id"])).fetchone()
    if not rec:
        abort(404)
    photos = db.execute("SELECT filename FROM photos WHERE record_id = ?", (record_id,)).fetchall()
    for p in photos:
        fpath = os.path.join(app.config["UPLOAD_FOLDER"], p["filename"])
        if os.path.exists(fpath):
            try: os.remove(fpath)
            except Exception: pass
    db.execute("DELETE FROM photos WHERE record_id = ?", (record_id,))
    db.execute("DELETE FROM records WHERE id = ?", (record_id,))
    db.commit()
    flash("Registro apagado.", "info")
    return redirect(url_for("dashboard"))

# ===== ADMIN =====
@app.route("/admin")
@admin_required
def admin_home():
    return render_template("admin_home.html")

@app.route("/admin/users", methods=["GET", "POST"])
@admin_required
def admin_users():
    db = get_db()
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        is_admin = 1 if request.form.get("is_admin") == "on" else 0
        if not username or not password:
            flash("Preencha usuário e senha.", "error")
            return redirect(url_for("admin_users"))
        try:
            pw_hash = generate_password_hash(password)
            db.execute("INSERT INTO users (username, password_hash, is_admin) VALUES (?, ?, ?)", (username, pw_hash, is_admin))
            db.commit()
            flash("Usuário criado com sucesso.", "success")
        except sqlite3.IntegrityError:
            flash("Nome de usuário já existe.", "error")
        return redirect(url_for("admin_users"))
    users = db.execute("SELECT id, username, is_admin FROM users ORDER BY username ASC").fetchall()
    return render_template("admin_users.html", users=users)

@app.route("/admin/users/<int:user_id>/toggle_admin", methods=["POST"])
@admin_required
def admin_toggle_admin(user_id):
    db = get_db()
    row = db.execute("SELECT is_admin FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        flash("Usuário não encontrado.", "error")
        return redirect(url_for("admin_users"))
    new_val = 0 if row["is_admin"] else 1
    db.execute("UPDATE users SET is_admin = ? WHERE id = ?", (new_val, user_id))
    db.commit()
    flash("Permissão atualizada.", "success")
    return redirect(url_for("admin_users"))

@app.route("/admin/users/<int:user_id>/reset", methods=["GET", "POST"])
@admin_required
def admin_reset_password(user_id):
    db = get_db()
    user = db.execute("SELECT id, username FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        flash("Usuário não encontrado.", "error")
        return redirect(url_for("admin_users"))
    if request.method == "POST":
        p1 = request.form.get("password", "").strip()
        p2 = request.form.get("password2", "").strip()
        if not p1 or not p2 or p1 != p2:
            flash("As senhas devem ser preenchidas e iguais.", "error")
            return redirect(url_for("admin_reset_password", user_id=user_id))
        pw_hash = generate_password_hash(p1)
        db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (pw_hash, user_id))
        db.commit()
        flash("Senha atualizada com sucesso.", "success")
        return redirect(url_for("admin_users"))
    return render_template("admin_reset_password.html", user=user)

@app.route("/admin/records")
@admin_required
def admin_records():
    user_id = request.args.get("user_id", type=int)
    db = get_db()
    if user_id:
        recs = db.execute(
            "SELECT r.id, r.device_name, r.fusion_count, r.created_at, u.username "
            "FROM records r JOIN users u ON u.id = r.user_id WHERE r.user_id = ? ORDER BY r.created_at DESC",
            (user_id,),
        ).fetchall()
    else:
        recs = db.execute(
            "SELECT r.id, r.device_name, r.fusion_count, r.created_at, u.username "
            "FROM records r JOIN users u ON u.id = r.user_id ORDER BY r.created_at DESC"
        ).fetchall()
    users = db.execute("SELECT id, username FROM users ORDER BY username ASC").fetchall()
    return render_template("admin_records.html", records=recs, users=users, selected_user_id=user_id)

@app.route("/admin/export.csv")
@admin_required
def admin_export_csv():
    user_id = request.args.get("user_id", type=int)
    db = get_db()
    if user_id:
        rows = db.execute(
            "SELECT r.id, u.username, r.device_name, r.fusion_count, r.created_at "
            "FROM records r JOIN users u ON u.id = r.user_id WHERE r.user_id = ? ORDER BY r.created_at DESC",
            (user_id,),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT r.id, u.username, r.device_name, r.fusion_count, r.created_at "
            "FROM records r JOIN users u ON u.id = r.user_id ORDER BY r.created_at DESC"
        ).fetchall()
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(["id", "username", "device_name", "fusion_count", "created_at"])
    for r in rows:
        writer.writerow([r["id"], r["username"], r["device_name"], r["fusion_count"], r["created_at"]])
    return Response(si.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=registros_splicing_admin.csv"})

# ===== Relatórios com filtros + gráficos + XLSX =====
from datetime import datetime

@app.route("/admin/reports", methods=["GET"])
@admin_required
def admin_reports():
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    db = get_db()
    rows = db.execute(
        f"SELECT r.id, u.username, r.device_name, r.fusion_count, r.created_at "
        f"FROM records r JOIN users u ON u.id = r.user_id {where_sql} ORDER BY r.created_at DESC",
        tuple(params),
    ).fetchall()

    total_fusions = db.execute(
        f"SELECT COALESCE(SUM(r.fusion_count), 0) as total FROM records r {where_sql}",
        tuple(params),
    ).fetchone()["total"] or 0

    users_summary = db.execute(
        f"SELECT u.id, u.username, "
        f"COUNT(DISTINCT r.device_name) AS devices, "
        f"COUNT(r.id) AS registros, "
        f"COALESCE(SUM(r.fusion_count), 0) AS fusoes "
        f"FROM records r JOIN users u ON u.id = r.user_id "
        f"{where_sql} "
        f"GROUP BY u.id, u.username "
        f"ORDER BY fusoes DESC, devices DESC",
        tuple(params),
    ).fetchall()

    devices = db.execute(
        f"SELECT r.device_name, COUNT(*) as registros, SUM(r.fusion_count) as fusoes "
        f"FROM records r {where_sql} GROUP BY r.device_name ORDER BY fusoes DESC, registros DESC",
        tuple(params),
    ).fetchall()

    users = db.execute("SELECT id, username FROM users ORDER BY username ASC").fetchall()

    return render_template(
        "admin_reports.html",
        rows=rows, users=users, selected_user_id=user_id,
        start=start_str, end=end_str,
        total_fusions=total_fusions, devices=devices, users_summary=users_summary
    )

@app.route("/admin/reports_data.json")
@admin_required
def admin_reports_data():
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    db = get_db()

    by_day = db.execute(
        f"SELECT date(r.created_at) as d, SUM(r.fusion_count) as total FROM records r {where_sql} GROUP BY date(r.created_at) ORDER BY d ASC",
        tuple(params),
    ).fetchall()
    by_day_list = [{"date": r["d"], "sum": r["total"] or 0} for r in by_day]

    by_device = db.execute(
        f"SELECT r.device_name, SUM(r.fusion_count) as total FROM records r {where_sql} GROUP BY r.device_name ORDER BY total DESC",
        tuple(params),
    ).fetchall()
    by_device_list = [{"device_name": r["device_name"], "sum": r["total"] or 0} for r in by_device]

    total_fusions = sum(item["sum"] for item in by_day_list)
    return {"by_day": by_day_list, "by_device": by_device_list, "total_fusions": total_fusions}

@app.route("/admin/reports.csv")
@admin_required
def admin_reports_csv():
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    db = get_db()
    rows = db.execute(
        f"SELECT r.id, u.username, r.device_name, r.fusion_count, r.created_at FROM records r JOIN users u ON u.id = r.user_id {where_sql} ORDER BY r.created_at DESC",
        tuple(params),
    ).fetchall()
    si = StringIO(); writer = csv.writer(si)
    writer.writerow(["id", "username", "device_name", "fusion_count", "created_at"])
    for r in rows:
        writer.writerow([r["id"], r["username"], r["device_name"], r["fusion_count"], r["created_at"]])
    return Response(si.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=relatorio_admin.csv"})

@app.route("/admin/reports_users.csv")
@admin_required
def admin_reports_users_csv():
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    db = get_db()
    rows = db.execute(
        f"SELECT u.id, u.username, COUNT(DISTINCT r.device_name) AS devices, COUNT(r.id) AS registros, COALESCE(SUM(r.fusion_count), 0) AS fusoes "
        f"FROM records r JOIN users u ON u.id = r.user_id {where_sql} GROUP BY u.id, u.username ORDER BY fusoes DESC, devices DESC",
        tuple(params),
    ).fetchall()

    si = StringIO(); writer = csv.writer(si)
    writer.writerow(["user_id", "username", "devices_distintos", "registros", "fusoes"])
    for r in rows:
        writer.writerow([r["id"], r["username"], r["devices"], r["registros"], r["fusoes"]])
    return Response(si.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=relatorio_por_usuario.csv"})

@app.route("/admin/reports.xlsx")
@admin_required
def admin_reports_xlsx():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str: clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:   clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:   clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    db = get_db()

    rows = db.execute(
        f"SELECT r.id, u.username, r.device_name, r.fusion_count, r.created_at FROM records r JOIN users u ON u.id = r.user_id {where_sql} ORDER BY r.created_at DESC",
        tuple(params),
    ).fetchall()
    devices = db.execute(
        f"SELECT r.device_name, COUNT(*) as registros, SUM(r.fusion_count) as fusoes FROM records r {where_sql} GROUP BY r.device_name ORDER BY fusoes DESC",
        tuple(params),
    ).fetchall()
    users_summary = db.execute(
        f"SELECT u.id, u.username, COUNT(DISTINCT r.device_name) AS devices, COUNT(r.id) AS registros, COALESCE(SUM(r.fusion_count), 0) AS fusoes "
        f"FROM records r JOIN users u ON u.id = r.user_id {where_sql} GROUP BY u.id, u.username ORDER BY fusoes DESC, devices DESC",
        tuple(params),
    ).fetchall()
    by_day = db.execute(
        f"SELECT date(r.created_at) as d, SUM(r.fusion_count) as total FROM records r {where_sql} GROUP BY date(r.created_at) ORDER BY d ASC",
        tuple(params),
    ).fetchall()

    wb = Workbook()
    ws1 = wb.active; ws1.title = "Detalhes"
    ws1.append(["id", "username", "device_name", "fusion_count", "created_at"])
    for r in rows:
        ws1.append([r["id"], r["username"], r["device_name"], r["fusion_count"], r["created_at"]])

    ws2 = wb.create_sheet("Dispositivos")
    ws2.append(["device_name", "registros", "fusoes"])
    for d in devices:
        ws2.append([d["device_name"], d["registros"], d["fusoes"]])

    ws3 = wb.create_sheet("Por Dia")
    ws3.append(["date", "fusoes"])
    total = 0
    for r in by_day:
        v = r["total"] or 0
        total += v
        ws3.append([r["d"], v])
    ws3.append([]); ws3.append(["TOTAL", total])

    ws4 = wb.create_sheet("Por Usuário")
    ws4.append(["user_id", "username", "devices_distintos", "registros", "fusoes"])
    for r in users_summary:
        ws4.append([r["id"], r["username"], r["devices"], r["registros"], r["fusoes"]])

    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col[0].column)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return Response(bio.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=relatorio_admin.xlsx"})

# ===== Download de fotos em ZIP por dispositivo =====
@app.route("/admin/photos", methods=["GET", "POST"])
@admin_required
def admin_photos():
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    db = get_db()
    devices = db.execute(
        f"""
        SELECT r.device_name,
               COUNT(DISTINCT r.id) as registros,
               COUNT(p.id) as fotos
        FROM records r
        LEFT JOIN photos p ON p.record_id = r.id
        {where_sql}
        GROUP BY r.device_name
        ORDER BY fotos DESC, registros DESC
        """,
        tuple(params)
    ).fetchall()

    users = db.execute("SELECT id, username FROM users ORDER BY username ASC").fetchall()
    return render_template("admin_photos.html", devices=devices, users=users, selected_user_id=user_id, start=start_str, end=end_str)

@app.route("/admin/photos.zip", methods=["POST"])
@admin_required
def admin_photos_zip():
    start_str = request.form.get("start", "").strip()
    end_str = request.form.get("end", "").strip()
    user_id = request.form.get("user_id", type=int)
    selected = request.form.getlist("devices")

    if not selected:
        flash("Selecione pelo menos um dispositivo.", "error")
        return redirect(url_for("admin_photos", start=start_str, end=end_str, user_id=user_id))

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    in_clause = " OR ".join(["r.device_name = ?"] * len(selected))
    clauses.append(f"({in_clause})"); params.extend(selected)

    where_sql = "WHERE " + " AND ".join(clauses)

    db = get_db()
    rows = db.execute(
        f"""
        SELECT r.id as record_id, r.device_name, u.username, p.filename
        FROM records r
        JOIN users u ON u.id = r.user_id
        JOIN photos p ON p.record_id = r.id
        {where_sql}
        ORDER BY r.device_name ASC, r.id ASC
        """,
        tuple(params)
    ).fetchall()

    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as z:
        base_upload = app.config.get("UPLOAD_FOLDER")
        for row in rows:
            device = row["device_name"]
            rec_id = row["record_id"]
            filename = row["filename"]
            fpath = os.path.join(base_upload, filename)
            if os.path.isfile(fpath):
                arcname = f"{device}/record_{rec_id}/{filename}"
                z.write(fpath, arcname)
    bio.seek(0)
    fname = "fotos_filtradas.zip"
    return Response(
        bio.getvalue(),
        mimetype="application/zip",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

# ===== Export do usuário (pessoal) =====
@app.route("/export.csv")
@login_required
def export_csv():
    db = get_db()
    rows = db.execute(
        "SELECT id, device_name, fusion_count, created_at FROM records WHERE user_id = ? ORDER BY created_at DESC",
        (session["user_id"],),
    ).fetchall()
    si = StringIO(); writer = csv.writer(si)
    writer.writerow(["id", "device_name", "fusion_count", "created_at", "photo_urls"])
    for r in rows:
        photos = db.execute("SELECT filename FROM photos WHERE record_id = ?", (r["id"],)).fetchall()
        host = request.host_url.rstrip("/")
        urls = [f"{host}{url_for('uploaded_file', filename=p['filename'])}" for p in photos]
        writer.writerow([r["id"], r["device_name"], r["fusion_count"], r["created_at"], " | ".join(urls)])
    return Response(si.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=registros_splicing.csv"})

# ===== Rota de emergência para resetar senha do admin =====
@app.route("/force_reset_admin")
def force_reset_admin():
    # Controle por variáveis de ambiente
    if os.environ.get("FORCE_RESET_ADMIN", "0") != "1":
        return "Desativado", 403
    token = request.args.get("token", "")
    expected = os.environ.get("RESET_ADMIN_TOKEN", "")
    if not token or token != expected:
        return "Token inválido", 403
    new_pw = os.environ.get("NEW_ADMIN_PASSWORD", "nova123")
    db = get_db()
    # Reseta a senha do primeiro admin encontrado
    row = db.execute("SELECT id FROM users WHERE is_admin = 1 ORDER BY id ASC LIMIT 1").fetchone()
    if not row:
        return "Nenhum admin encontrado", 404
    db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_pw), row["id"]))
    db.commit()
    return f"Senha do admin (id={row['id']}) resetada para: {new_pw}"

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
